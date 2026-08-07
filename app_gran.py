"""
Hympyr Énergies — Cockpit granulés de bois
==========================================

4 onglets :
  1. Simulateur      — prix HT/TTC + frais de livraison (distance routière réelle)
  2. Grille          — grille tarifaire datée, éditable en HT ou TTC, exportable
  3. Données         — import de grilles historiques et de tarifs concurrents
  4. Analyse         — grilles de lecture des prix (HT ou TTC) + rapports

Chaîne technique distance :
  Géocodage CP+ville -> API Adresse (BAN, gratuit, sans clé)
  Distance routière  -> OSRM public (gratuit, sans clé)
  Repli              -> orthodromique majorée, puis saisie manuelle

Règle d'arrondi TVA : au moment de la validation de la grille, les deux
montants (HT et TTC) sont figés et stockés. Aucun recalcul en boucle, ce qui
évite les dérives d'un centime au retour de conversion.

Persistance : session uniquement (Streamlit Cloud a un système de fichiers
éphémère). L'archivage se fait par export/import Excel.

Usage interne équipe commerciale.
"""

import io
import math
import os
import re
import unicodedata
from datetime import date, datetime

import matplotlib

matplotlib.use("Agg")
import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
import requests
import streamlit as st

# ══════════════════════════════════════════════════════════════════════════════
# CONFIGURATION MÉTIER — tout ce qui change se modifie ICI
# ══════════════════════════════════════════════════════════════════════════════

ADRESSE_DEPART = "490 route de Toulouse, 81370 Saint-Sulpice-la-Pointe"
TELEPHONE = "05 61 70 03 27"
SOCIETE = "Hympyr Énergies"

# ⚠️ TAUX DE TVA — À VALIDER PAR LA COMPTABILITÉ AVANT USAGE COMMERCIAL.
# Valeurs de départ, modifiables dans l'onglet « Grille tarifaire ».
# Le taux applicable dépend du produit, de l'usage (domestique / professionnel)
# et du statut du client. Les frais de port suivent en principe le taux de la
# prestation principale, mais ce point doit être confirmé.
TVA_MARCHANDISE_DEFAUT = 0.10
TVA_LIVRAISON_DEFAUT = 0.10

# Identité visuelle (centralisée : une seule ligne à changer si la charte évolue)
VERT_FONCE = "#005727"
VERT_VIF = "#14B02F"
VERT_CLAIR = "#f0faf3"
GRIS = "#666666"
ROUGE = "#e53935"
ORANGE = "#f0a202"

# Grille livraison palettes, exprimée en TTC (bornes CONTINUES : pas de trou)
ZONES_TTC_DEFAUT = [
    {"Zone": "Zone 1", "Jusqu'à (km)": 10.0, "72 h TTC": 49.0, "15 j TTC": 29.0},
    {"Zone": "Zone 2", "Jusqu'à (km)": 20.0, "72 h TTC": 49.0, "15 j TTC": 35.0},
    {"Zone": "Zone 3", "Jusqu'à (km)": 30.0, "72 h TTC": 59.0, "15 j TTC": 45.0},
    {"Zone": "Zone 4", "Jusqu'à (km)": 53.0, "72 h TTC": 79.0, "15 j TTC": 59.0},
    {"Zone": "Zone 5", "Jusqu'à (km)": 9999.0, "72 h TTC": 95.0, "15 j TTC": 69.0},
]

VRAC_FRANCHISE_KM_DEFAUT = 35.0
VRAC_PRIX_KM_TTC_DEFAUT = 1.40  # € TTC / km au-delà de la franchise

PALETTES_TTC_DEFAUT = {
    "Piveteau": 452.87,
    "Granulés de nos régions": 431.24,
}

VRAC_TTC_DEFAUT = {  # tonnage -> € TTC / tonne
    2: 437.0, 3: 420.0, 4: 415.0, 5: 409.0,
    6: 404.0, 7: 404.0, 8: 399.0, 9: 399.0,
}

SEUIL_ALERTE_KM = 80.0
ARRONDI_KM = "entier"  # "entier" | "superieur" | "aucun"
COEF_VOL_OISEAU = 1.30

# Schéma canonique de l'historique de prix
COLONNES = [
    "date",            # date d'effet du tarif
    "acteur",          # "Hympyr" ou nom du concurrent
    "type_acteur",     # "Interne" | "Concurrent"
    "conditionnement", # "Palette" | "Vrac"
    "produit",         # libellé produit
    "palier_t",        # palier de tonnage (vrac), sinon vide
    "unite",           # "palette" | "tonne"
    "prix_ht",         # € HT
    "taux_tva",        # taux décimal (0.10 = 10 %)
    "prix_ttc",        # € TTC
]

st.set_page_config(
    page_title="Hympyr – Cockpit granulés",
    page_icon="🌿",
    layout="wide",
)

# ══════════════════════════════════════════════════════════════════════════════
# ===== LOGIQUE PURE — DÉBUT (testable hors Streamlit) =========================
# ══════════════════════════════════════════════════════════════════════════════

def eur(v) -> str:
    """Montant en euros, format français."""
    if v is None or (isinstance(v, float) and math.isnan(v)):
        return "—"
    s = f"{v:,.2f}".replace(",", "@").replace(".", ",").replace("@", "\u202f")
    return f"{s} €"


def dec(v, n: int = 2) -> str:
    if v is None or (isinstance(v, float) and math.isnan(v)):
        return "—"
    return f"{v:.{n}f}".replace(".", ",")


def km_fmt(v: float) -> str:
    return dec(v, 1) + " km"


def pct(v, signe: bool = True) -> str:
    if v is None or (isinstance(v, float) and math.isnan(v)):
        return "—"
    s = "+" if (signe and v > 0) else ""
    return s + dec(v, 1) + " %"


def taux_fmt(t) -> str:
    """0.10 -> '10 %'"""
    if t is None or (isinstance(t, float) and math.isnan(t)):
        return "—"
    return dec(float(t) * 100, 1).rstrip("0").rstrip(",") + " %"


# ── Conversions TVA ───────────────────────────────────────────────────────────

def ht_vers_ttc(ht, taux) -> float:
    if ht is None or (isinstance(ht, float) and math.isnan(ht)):
        return float("nan")
    return round(float(ht) * (1.0 + float(taux)), 2)


def ttc_vers_ht(ttc, taux) -> float:
    if ttc is None or (isinstance(ttc, float) and math.isnan(ttc)):
        return float("nan")
    return round(float(ttc) / (1.0 + float(taux)), 2)


def normaliser_taux(v) -> float:
    """Accepte 10, '10 %', 0.10, '0,1' -> 0.10. NaN si illisible."""
    if v is None:
        return float("nan")
    if isinstance(v, str):
        v = v.replace("%", "").replace(",", ".").strip()
        if not v:
            return float("nan")
    try:
        x = float(v)
    except (TypeError, ValueError):
        return float("nan")
    if math.isnan(x):
        return float("nan")
    return x / 100.0 if x > 1.0 else x


def paire_prix(valeur, taux, base: str):
    """
    Depuis un montant saisi dans une base donnée, renvoie (ht, ttc).
    Les deux valeurs sont figées ici : aucun recalcul ultérieur, donc aucune
    dérive d'arrondi au retour de conversion.
    """
    if valeur is None or (isinstance(valeur, float) and math.isnan(valeur)):
        return (float("nan"), float("nan"))
    v = float(valeur)
    if str(base).upper() == "HT":
        return (round(v, 2), ht_vers_ttc(v, taux))
    return (ttc_vers_ht(v, taux), round(v, 2))


def arrondir_km(km: float) -> float:
    if ARRONDI_KM == "entier":
        return float(round(km))
    if ARRONDI_KM == "superieur":
        return float(math.ceil(km))
    return km


def haversine(lat1, lon1, lat2, lon2) -> float:
    r = 6371.0
    p1, p2 = math.radians(lat1), math.radians(lat2)
    dp, dl = math.radians(lat2 - lat1), math.radians(lon2 - lon1)
    a = math.sin(dp / 2) ** 2 + math.cos(p1) * math.cos(p2) * math.sin(dl / 2) ** 2
    return 2 * r * math.asin(math.sqrt(a))


def normaliser(txt) -> str:
    """Minuscules, sans accents ni ponctuation — pour l'appariement de colonnes."""
    t = unicodedata.normalize("NFKD", str(txt))
    t = "".join(c for c in t if not unicodedata.combining(c))
    return re.sub(r"[^a-z0-9]+", "_", t.lower()).strip("_")


# ── Tarification ──────────────────────────────────────────────────────────────

def libelle_zone(precedent: float, borne: float) -> str:
    if precedent == 0:
        return f"≤ {dec(borne, 0)} km"
    if borne < 9999:
        return f"> {dec(precedent, 0)} à {dec(borne, 0)} km"
    return f"> {dec(precedent, 0)} km"


def zone_palette(km: float, zones: list):
    """-> (nom_zone, libellé, dict_zone). Bornes continues, tri croissant."""
    tri = sorted(zones, key=lambda z: float(z["Jusqu'à (km)"]))
    precedent = 0.0
    for z in tri:
        borne = float(z["Jusqu'à (km)"])
        if km <= borne:
            return z["Zone"], libelle_zone(precedent, borne), z
        precedent = borne
    d = tri[-1]
    return d["Zone"], f"> {dec(precedent, 0)} km", d


def frais_vrac(km: float, franchise: float, prix_km: float) -> float:
    """
    Franchise puis tarif linéaire au km. Le résultat est exprimé dans la même
    base (HT ou TTC) que `prix_km`.
    """
    return round(max(0.0, float(km) - float(franchise)) * float(prix_km), 2)


def cout_livre(prix_produit: float, frais: float, tonnes: float) -> float:
    """Coût rendu à la tonne — le seul comparable réellement honnête."""
    if not tonnes or tonnes <= 0:
        return float("nan")
    return (prix_produit + frais) / tonnes


# ── Modèle de données ─────────────────────────────────────────────────────────

def df_vide() -> pd.DataFrame:
    df = pd.DataFrame({c: pd.Series(dtype="object") for c in COLONNES})
    for c in ("prix_ht", "prix_ttc", "palier_t", "taux_tva"):
        df[c] = pd.Series(dtype="float64")
    return df


def reference(row) -> str:
    """Clé produit unifiée : 'Piveteau', 'Vrac 4 T', ..."""
    if str(row["conditionnement"]).strip().lower().startswith("vrac"):
        p = row.get("palier_t")
        try:
            if p is None or (isinstance(p, float) and math.isnan(p)):
                return "Vrac"
            return f"Vrac {int(float(p))} T"
        except (TypeError, ValueError):
            return "Vrac"
    return str(row["produit"]).strip()


def completer_prix(d: pd.DataFrame, taux_defaut: float) -> pd.DataFrame:
    """Complète la colonne manquante (HT ou TTC) à partir du taux disponible."""
    d = d.copy()
    d["taux_tva"] = d["taux_tva"].apply(normaliser_taux)
    d["taux_tva"] = d["taux_tva"].fillna(float(taux_defaut))

    manque_ht = d["prix_ht"].isna() & d["prix_ttc"].notna()
    if manque_ht.any():
        d.loc[manque_ht, "prix_ht"] = [
            ttc_vers_ht(t, x) for t, x in
            zip(d.loc[manque_ht, "prix_ttc"], d.loc[manque_ht, "taux_tva"])]

    manque_ttc = d["prix_ttc"].isna() & d["prix_ht"].notna()
    if manque_ttc.any():
        d.loc[manque_ttc, "prix_ttc"] = [
            ht_vers_ttc(h, x) for h, x in
            zip(d.loc[manque_ttc, "prix_ht"], d.loc[manque_ttc, "taux_tva"])]
    return d


def enrichir(df: pd.DataFrame, taux_defaut: float = TVA_MARCHANDISE_DEFAUT) -> pd.DataFrame:
    """Ajoute 'reference', normalise les types, complète HT/TTC."""
    vide = df_vide()
    vide["reference"] = pd.Series(dtype="object")
    if df is None or df.empty:
        return vide

    d = df.copy()
    for c in COLONNES:
        if c not in d.columns:
            d[c] = np.nan
    d["date"] = pd.to_datetime(d["date"], errors="coerce")
    for c in ("prix_ht", "prix_ttc", "palier_t"):
        d[c] = pd.to_numeric(d[c], errors="coerce")

    d = d.dropna(subset=["date"])
    d = d[d["prix_ht"].notna() | d["prix_ttc"].notna()]
    if d.empty:
        return vide

    d = completer_prix(d, taux_defaut)
    d["reference"] = d.apply(reference, axis=1)
    return d


def grille_vers_lignes(date_effet, palettes: dict, vrac: dict, taux: float,
                       acteur: str = "Hympyr") -> pd.DataFrame:
    """
    Convertit une grille tarifaire en lignes canoniques.
    `palettes` et `vrac` : {clé: {"ht": x, "ttc": y}}
    """
    lignes = []
    for nom, p in palettes.items():
        lignes.append({
            "date": pd.Timestamp(date_effet), "acteur": acteur,
            "type_acteur": "Interne", "conditionnement": "Palette",
            "produit": nom, "palier_t": np.nan, "unite": "palette",
            "prix_ht": float(p["ht"]), "taux_tva": float(taux),
            "prix_ttc": float(p["ttc"]),
        })
    for t, p in sorted(vrac.items()):
        lignes.append({
            "date": pd.Timestamp(date_effet), "acteur": acteur,
            "type_acteur": "Interne", "conditionnement": "Vrac",
            "produit": "Granulés vrac", "palier_t": float(t), "unite": "tonne",
            "prix_ht": float(p["ht"]), "taux_tva": float(taux),
            "prix_ttc": float(p["ttc"]),
        })
    return pd.DataFrame(lignes, columns=COLONNES)


ALIAS = {
    "date": ["date", "date_effet", "date_d_effet", "periode", "mois", "jour"],
    "acteur": ["acteur", "fournisseur", "concurrent", "enseigne", "societe",
               "entreprise", "marque"],
    "type_acteur": ["type_acteur", "type", "categorie", "nature"],
    "conditionnement": ["conditionnement", "format", "type_produit", "packaging"],
    "produit": ["produit", "libelle", "designation", "article", "reference"],
    "palier_t": ["palier_t", "palier", "tonnage", "tonnes", "quantite_t", "qte"],
    "unite": ["unite", "unit"],
    "prix_ht": ["prix_ht", "prix_hors_taxe", "prix_hors_taxes", "ht", "montant_ht",
                "tarif_ht", "prix_unitaire_ht", "prix_tonne_ht", "prix_palette_ht"],
    "taux_tva": ["taux_tva", "tva", "taux", "taux_de_tva", "tva_pct"],
    "prix_ttc": ["prix_ttc", "prix", "tarif", "montant", "montant_ttc", "ttc",
                 "prix_ttc_eur", "prix_unitaire", "prix_unitaire_ttc",
                 "prix_tonne", "prix_palette"],
}


def mapper_colonnes(cols) -> dict:
    """Apparie les colonnes d'un fichier importé au schéma canonique."""
    norm = {}
    for c in cols:
        norm.setdefault(normaliser(c), c)
    mapping = {}
    for cible, alias in ALIAS.items():
        for a in alias:
            if a in norm and norm[a] not in mapping.values():
                mapping[cible] = norm[a]
                break
    return mapping


def _num_fr(serie: pd.Series) -> pd.Series:
    """Nettoie une colonne de montants au format français."""
    return pd.to_numeric(
        serie.astype(str)
        .str.replace("\u202f", "", regex=False)
        .str.replace("\u00a0", "", regex=False)
        .str.replace(" ", "", regex=False)
        .str.replace("€", "", regex=False)
        .str.replace(",", ".", regex=False),
        errors="coerce")


def valider_import(brut: pd.DataFrame, nom_fichier: str = "",
                   taux_defaut: float = TVA_MARCHANDISE_DEFAUT):
    """
    Normalise un fichier importé vers le schéma canonique.
    -> (df_valide, rejets, avertissements)
    """
    avertissements, rejets = [], []
    mapping = mapper_colonnes(brut.columns)

    manquantes = [c for c in ("date", "acteur") if c not in mapping]
    if "prix_ht" not in mapping and "prix_ttc" not in mapping:
        manquantes.append("prix HT ou prix TTC")
    if manquantes:
        rejets.append(f"{nom_fichier} : colonnes obligatoires introuvables "
                      f"({', '.join(manquantes)}). Fichier ignoré.")
        return df_vide(), rejets, avertissements

    d = pd.DataFrame(index=brut.index)
    for cible in COLONNES:
        d[cible] = brut[mapping[cible]] if cible in mapping else np.nan

    d["date"] = pd.to_datetime(d["date"], errors="coerce", dayfirst=True)
    for c in ("prix_ht", "prix_ttc"):
        d[c] = _num_fr(d[c]) if c in mapping else np.nan
    d["palier_t"] = pd.to_numeric(d["palier_t"], errors="coerce")
    d["taux_tva"] = d["taux_tva"].apply(normaliser_taux)

    n0 = len(d)
    sans_prix = d["prix_ht"].isna() & d["prix_ttc"].isna()
    invalides = d["date"].isna() | sans_prix
    if invalides.any():
        rejets.append(f"{nom_fichier} : {int(invalides.sum())} ligne(s) sur {n0} "
                      "écartée(s) (date ou prix illisible).")
    d = d[~invalides].copy()
    if d.empty:
        return df_vide(), rejets, avertissements

    if "taux_tva" not in mapping or d["taux_tva"].isna().all():
        avertissements.append(
            f"{nom_fichier} : taux de TVA absent, {taux_fmt(taux_defaut)} appliqué "
            "par défaut pour compléter la colonne manquante.")
    elif d["taux_tva"].isna().any():
        avertissements.append(
            f"{nom_fichier} : {int(d['taux_tva'].isna().sum())} ligne(s) sans taux "
            f"de TVA, complétées à {taux_fmt(taux_defaut)}.")
    d = completer_prix(d, taux_defaut)

    if "conditionnement" not in mapping:
        d["conditionnement"] = np.where(d["palier_t"].notna(), "Vrac", "Palette")
        avertissements.append(
            f"{nom_fichier} : colonne « conditionnement » absente, déduite du palier.")
    d["conditionnement"] = (d["conditionnement"].astype(str).str.strip()
                            .str.capitalize().replace({"Nan": "Palette", "": "Palette"}))

    d["acteur"] = d["acteur"].astype(str).str.strip()

    if "type_acteur" not in mapping:
        d["type_acteur"] = np.where(
            d["acteur"].str.lower().str.contains("hympyr"), "Interne", "Concurrent")
        avertissements.append(f"{nom_fichier} : colonne « type_acteur » absente, "
                              "déduite du nom de l'acteur.")

    if "produit" not in mapping:
        d["produit"] = np.where(d["conditionnement"] == "Vrac",
                                "Granulés vrac", "Palette")
    if "unite" not in mapping:
        d["unite"] = np.where(d["conditionnement"] == "Vrac", "tonne", "palette")

    return d[COLONNES].reset_index(drop=True), rejets, avertissements


def dedoublonner(df: pd.DataFrame,
                 taux_defaut: float = TVA_MARCHANDISE_DEFAUT) -> pd.DataFrame:
    """Une seule ligne par (date, acteur, référence) — la dernière importée gagne."""
    if df is None or df.empty:
        return df_vide()
    d = enrichir(df, taux_defaut)
    if d.empty:
        return df_vide()
    d = d.drop_duplicates(subset=["date", "acteur", "reference"], keep="last")
    return d[COLONNES].reset_index(drop=True)


# ── Grilles de lecture ────────────────────────────────────────────────────────
# Toutes acceptent `colonne` = "prix_ttc" (défaut) ou "prix_ht".

def matrice_dernier_prix(df: pd.DataFrame, colonne: str = "prix_ttc") -> pd.DataFrame:
    """Dernier prix connu par référence × acteur."""
    d = enrichir(df)
    if d.empty:
        return pd.DataFrame()
    d = d.sort_values("date")
    dern = d.groupby(["reference", "acteur"], as_index=False).last()
    return dern.pivot(index="reference", columns="acteur", values=colonne)


def positionnement(df: pd.DataFrame, interne: str = "Hympyr",
                   colonne: str = "prix_ttc") -> pd.DataFrame:
    """
    Positionnement sur le dernier prix connu.
    Indice base 100 = moyenne des concurrents. < 100 : Hympyr moins cher.
    """
    m = matrice_dernier_prix(df, colonne)
    if m.empty or interne not in m.columns:
        return pd.DataFrame()
    concurrents = [c for c in m.columns if c != interne]
    if not concurrents:
        return pd.DataFrame()

    out = pd.DataFrame(index=m.index)
    out["Hympyr"] = m[interne]
    out["Moyenne marché"] = m[concurrents].mean(axis=1)
    out["Mini marché"] = m[concurrents].min(axis=1)
    out["Maxi marché"] = m[concurrents].max(axis=1)
    out["Écart € vs moyenne"] = out["Hympyr"] - out["Moyenne marché"]
    out["Écart % vs moyenne"] = (out["Hympyr"] / out["Moyenne marché"] - 1) * 100
    out["Indice (base 100)"] = out["Hympyr"] / out["Moyenne marché"] * 100

    rangs = []
    for i in m.index:
        serie = m.loc[i].dropna()
        rangs.append(int(serie.rank(method="min")[interne])
                     if interne in serie.index else np.nan)
    out["Rang"] = rangs
    out["Nb offres comparées"] = m.notna().sum(axis=1)
    return out.dropna(subset=["Hympyr", "Moyenne marché"])


def volatilite(df: pd.DataFrame, colonne: str = "prix_ttc") -> pd.DataFrame:
    """Dispersion des prix par référence sur toute la période."""
    d = enrichir(df)
    if d.empty:
        return pd.DataFrame()
    g = d.groupby("reference")[colonne]
    out = pd.DataFrame({
        "Nb relevés": g.count(), "Prix mini": g.min(), "Prix maxi": g.max(),
        "Prix moyen": g.mean(), "Écart-type": g.std(),
    })
    out["Amplitude %"] = (out["Prix maxi"] / out["Prix mini"] - 1) * 100
    return out.sort_values("Amplitude %", ascending=False)


def evolution_periode(df: pd.DataFrame, acteur: str = "Hympyr",
                      colonne: str = "prix_ttc") -> pd.DataFrame:
    """Variation du premier au dernier relevé, par référence, pour un acteur."""
    d = enrichir(df)
    if d.empty:
        return pd.DataFrame()
    d = d[d["acteur"] == acteur]
    if d.empty:
        return pd.DataFrame()
    d = d.sort_values("date")
    prem, dern = d.groupby("reference").first(), d.groupby("reference").last()
    out = pd.DataFrame({
        "Première date": prem["date"].dt.date,
        "Prix initial": prem[colonne],
        "Dernière date": dern["date"].dt.date,
        "Prix actuel": dern[colonne],
    })
    out["Variation €"] = out["Prix actuel"] - out["Prix initial"]
    out["Variation %"] = (out["Prix actuel"] / out["Prix initial"] - 1) * 100
    return out.sort_values("Variation %", ascending=False)


def degressivite(df: pd.DataFrame, colonne: str = "prix_ttc") -> pd.DataFrame:
    """Prix/tonne par palier et par acteur (dernier prix connu)."""
    d = enrichir(df)
    if d.empty:
        return pd.DataFrame()
    d = d[d["conditionnement"].astype(str).str.lower().str.startswith("vrac")
          & d["palier_t"].notna()]
    if d.empty:
        return pd.DataFrame()
    d = d.sort_values("date").groupby(["acteur", "palier_t"], as_index=False).last()
    return d.pivot(index="palier_t", columns="acteur", values=colonne).sort_index()


def coherence_bases(df: pd.DataFrame) -> pd.DataFrame:
    """
    Contrôle qualité sur les taux de TVA. Deux anomalies détectées :
      - taux hétérogènes au sein d'un même acteur (mélange de sources) ;
      - taux divergent du taux dominant du jeu de données (saisie douteuse).
    Aucun contrôle ne peut détecter un prix TTC saisi comme HT : seule la
    rigueur de saisie protège de ce cas.
    """
    d = enrichir(df)
    if d.empty:
        return pd.DataFrame()
    g = d.groupby("acteur")["taux_tva"]
    out = pd.DataFrame({"Taux mini": g.min(), "Taux maxi": g.max(),
                        "Nb relevés": g.count()})
    out["Taux hétérogène"] = out["Taux mini"] != out["Taux maxi"]

    modes = d["taux_tva"].mode()
    dominant = float(modes.iloc[0]) if not modes.empty else float("nan")
    out["Taux dominant"] = dominant
    out["Divergent"] = (out["Taux mini"] != dominant) | (out["Taux maxi"] != dominant)
    out["Anomalie"] = out["Taux hétérogène"] | out["Divergent"]
    return out


# ── Graphiques ────────────────────────────────────────────────────────────────

def _style_axes(ax, titre="", ylab=""):
    ax.set_title(titre, fontsize=11, fontweight="bold", color=VERT_FONCE, pad=12)
    ax.set_ylabel(ylab, fontsize=9)
    ax.grid(True, alpha=0.25, linestyle="--", linewidth=0.7)
    ax.spines["top"].set_visible(False)
    ax.spines["right"].set_visible(False)
    ax.tick_params(labelsize=8)


def _base_lbl(colonne: str) -> str:
    return "HT" if colonne == "prix_ht" else "TTC"


def fig_evolution(df: pd.DataFrame, ref: str, colonne: str = "prix_ttc"):
    d = enrichir(df)
    fig, ax = plt.subplots(figsize=(8, 3.6), dpi=140)
    b = _base_lbl(colonne)
    d = d[d["reference"] == ref].sort_values("date") if not d.empty else d
    if d.empty:
        ax.text(0.5, 0.5, "Aucune donnée", ha="center", va="center")
        _style_axes(ax, f"Évolution du prix — {ref}", f"€ {b}")
        return fig
    for acteur, sub in d.groupby("acteur"):
        interne = str(acteur).lower().startswith("hympyr")
        ax.plot(sub["date"], sub[colonne], marker="o", markersize=4,
                linewidth=2.4 if interne else 1.3,
                color=VERT_FONCE if interne else None,
                zorder=5 if interne else 2, label=acteur)
    _style_axes(ax, f"Évolution du prix — {ref}", f"€ {b}")
    ax.legend(fontsize=7.5, frameon=False, ncol=3)
    fig.autofmt_xdate(rotation=30)
    fig.tight_layout()
    return fig


def fig_vs_marche(df: pd.DataFrame, ref: str, interne: str = "Hympyr",
                  colonne: str = "prix_ttc"):
    """Hympyr vs moyenne marché, avec bande mini-maxi des concurrents."""
    d = enrichir(df)
    fig, ax = plt.subplots(figsize=(8, 3.6), dpi=140)
    b = _base_lbl(colonne)
    d = d[d["reference"] == ref] if not d.empty else d
    if d.empty:
        ax.text(0.5, 0.5, "Aucune donnée", ha="center", va="center")
        _style_axes(ax, f"Positionnement vs marché — {ref}", f"€ {b}")
        return fig

    conc = d[d["acteur"] != interne]
    if not conc.empty:
        g = conc.groupby("date")[colonne]
        stats = pd.DataFrame({"moy": g.mean(), "mini": g.min(), "maxi": g.max()})
        ax.fill_between(stats.index, stats["mini"], stats["maxi"],
                        color=VERT_VIF, alpha=0.13, label="Fourchette marché")
        ax.plot(stats.index, stats["moy"], linestyle="--", linewidth=1.6,
                color=GRIS, label="Moyenne marché")

    mine = d[d["acteur"] == interne].sort_values("date")
    if not mine.empty:
        ax.plot(mine["date"], mine[colonne], marker="o", markersize=5,
                linewidth=2.6, color=VERT_FONCE, label=interne, zorder=5)

    _style_axes(ax, f"Positionnement vs marché — {ref}", f"€ {b}")
    ax.legend(fontsize=7.5, frameon=False)
    fig.autofmt_xdate(rotation=30)
    fig.tight_layout()
    return fig


def fig_degressivite(df: pd.DataFrame, colonne: str = "prix_ttc"):
    piv = degressivite(df, colonne)
    fig, ax = plt.subplots(figsize=(8, 3.6), dpi=140)
    b = _base_lbl(colonne)
    if piv.empty:
        ax.text(0.5, 0.5, "Aucune donnée vrac", ha="center", va="center")
        _style_axes(ax, "Dégressivité vrac", f"€ {b} / tonne")
        return fig
    for acteur in piv.columns:
        interne = str(acteur).lower().startswith("hympyr")
        ax.plot(piv.index, piv[acteur], marker="o", markersize=4,
                linewidth=2.4 if interne else 1.3,
                color=VERT_FONCE if interne else None,
                zorder=5 if interne else 2, label=acteur)
    _style_axes(ax, "Dégressivité vrac — prix à la tonne par palier",
                f"€ {b} / tonne")
    ax.set_xlabel("Tonnage commandé", fontsize=9)
    ax.legend(fontsize=7.5, frameon=False, ncol=3)
    fig.tight_layout()
    return fig


def fig_positionnement(pos: pd.DataFrame, base: str = "TTC"):
    fig, ax = plt.subplots(figsize=(8, 3.8), dpi=140)
    if pos is None or pos.empty:
        ax.text(0.5, 0.5, "Aucun concurrent renseigné", ha="center", va="center")
        _style_axes(ax, "Écart vs moyenne marché", "")
        return fig
    d = pos.sort_values("Écart % vs moyenne")
    vals = d["Écart % vs moyenne"]
    couleurs = [ROUGE if v > 0 else VERT_VIF for v in vals]
    ax.barh([str(i) for i in d.index], vals, color=couleurs, height=0.6)
    ax.axvline(0, color=GRIS, linewidth=1)
    # Décalage proportionnel à l'échelle : sinon les étiquettes sortent du cadre
    # quand les écarts sont faibles.
    etendue = max(abs(float(vals.min())), abs(float(vals.max())), 0.1)
    marge = etendue * 0.06
    for i, v in enumerate(vals):
        ax.text(v + (marge if v >= 0 else -marge), i, pct(v), va="center",
                ha="left" if v >= 0 else "right", fontsize=7.5)
    ax.set_xlim(min(float(vals.min()) - etendue * 0.45, -etendue * 0.15),
                max(float(vals.max()) + etendue * 0.45, etendue * 0.15))
    _style_axes(ax, f"Écart Hympyr vs moyenne marché (dernier prix, {base})", "")
    ax.set_xlabel("% — négatif = Hympyr moins cher", fontsize=9)
    fig.tight_layout()
    return fig


def fig_to_png(fig) -> bytes:
    buf = io.BytesIO()
    fig.savefig(buf, format="png", bbox_inches="tight")
    plt.close(fig)
    return buf.getvalue()


# ── Exports ───────────────────────────────────────────────────────────────────

def modele_import() -> pd.DataFrame:
    """Gabarit d'import à distribuer (colonnes attendues + exemples)."""
    return pd.DataFrame([
        {"date": "01/09/2026", "acteur": "Hympyr", "type_acteur": "Interne",
         "conditionnement": "Palette", "produit": "Piveteau", "palier_t": "",
         "unite": "palette", "prix_ht": "411,70", "taux_tva": "10",
         "prix_ttc": "452,87"},
        {"date": "01/09/2026", "acteur": "Concurrent A", "type_acteur": "Concurrent",
         "conditionnement": "Vrac", "produit": "Granulés vrac", "palier_t": "5",
         "unite": "tonne", "prix_ht": "", "taux_tva": "10", "prix_ttc": "415,00"},
    ])


def grille_tableau(palettes: dict, vrac: dict, taux: float) -> pd.DataFrame:
    """Tableau produit prêt à exporter : HT, taux, TTC."""
    lignes = [{"Conditionnement": "Palette", "Produit": k, "Palier (T)": "",
               "Unité": "palette", "Prix HT": p["ht"], "TVA": taux_fmt(taux),
               "Prix TTC": p["ttc"]} for k, p in palettes.items()]
    lignes += [{"Conditionnement": "Vrac", "Produit": "Granulés vrac",
                "Palier (T)": t, "Unité": "tonne", "Prix HT": p["ht"],
                "TVA": taux_fmt(taux), "Prix TTC": p["ttc"]}
               for t, p in sorted(vrac.items())]
    return pd.DataFrame(lignes)


def libelles_zones(zones: list, taux: float) -> pd.DataFrame:
    zl, precedent = [], 0.0
    for z in sorted(zones, key=lambda x: float(x["Jusqu'à (km)"])):
        b = float(z["Jusqu'à (km)"])
        zl.append({"Zone": z["Zone"], "Distance": libelle_zone(precedent, b),
                   "72 h HT": float(z["72 h HT"]), "72 h TTC": float(z["72 h TTC"]),
                   "15 j HT": float(z["15 j HT"]), "15 j TTC": float(z["15 j TTC"]),
                   "TVA": taux_fmt(taux)})
        precedent = b
    return pd.DataFrame(zl)


def _formater_classeur(book):
    """Mise en forme : en-têtes, largeurs, formats monétaires, volets figés."""
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    fill = PatternFill("solid", fgColor=VERT_FONCE.lstrip("#"))
    font = Font(color="FFFFFF", bold=True, size=10)

    for ws in book.worksheets:
        for cell in ws[1]:
            if cell.value is not None:
                cell.fill = fill
                cell.font = font
                cell.alignment = Alignment(horizontal="center", vertical="center",
                                           wrap_text=True)
        ws.freeze_panes = "A2"
        for col in ws.columns:
            lettre = get_column_letter(col[0].column)
            longueur = max((len(str(c.value)) for c in col if c.value is not None),
                           default=8)
            ws.column_dimensions[lettre].width = min(max(longueur + 3, 11), 34)

        for i, e in enumerate([str(c.value or "") for c in ws[1]], start=1):
            n = normaliser(e)
            if "tva" in n or "taux" in n:
                continue  # ni € ni décimales forcées sur un taux
            fmt = None
            if any(k in n for k in ("prix", "montant", "moyenne", "mini", "maxi",
                                    "ecart", "ht", "ttc", "72_h", "15_j")):
                fmt = '# ##0.00 "\u20ac"'
            if "%" in e or "indice" in n:
                fmt = "0.0"
            if fmt:
                for row in ws.iter_rows(min_row=2, min_col=i, max_col=i):
                    for c in row:
                        c.number_format = fmt


def export_excel(date_grille, palettes: dict, vrac: dict, zones: list,
                 historique: pd.DataFrame, taux_march: float, taux_livr: float,
                 franchise: float, prix_km_ht: float, prix_km_ttc: float) -> bytes:
    """Classeur multi-onglets : grille datée + référentiel + données + analyses."""
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as xw:
        g = grille_tableau(palettes, vrac, taux_march)
        g.insert(0, "Date d'effet", pd.Timestamp(date_grille).date())
        g.to_excel(xw, sheet_name="Grille tarifaire", index=False)

        z = libelles_zones(zones, taux_livr)
        z.to_excel(xw, sheet_name="Livraison", index=False)
        pd.DataFrame([
            {"Règle vrac": "Franchise (km)", "Valeur": franchise},
            {"Règle vrac": "€ HT / km au-delà", "Valeur": prix_km_ht},
            {"Règle vrac": "€ TTC / km au-delà", "Valeur": prix_km_ttc},
            {"Règle vrac": "TVA livraison", "Valeur": taux_fmt(taux_livr)},
        ]).to_excel(xw, sheet_name="Livraison", index=False, startrow=len(z) + 3)

        h = enrichir(historique, taux_march)
        if not h.empty:
            hh = h.copy()
            hh["date"] = hh["date"].dt.date
            hh.to_excel(xw, sheet_name="Données brutes", index=False)

            for suffixe, col in (("TTC", "prix_ttc"), ("HT", "prix_ht")):
                for nom, tab in (
                    (f"Matrice {suffixe}", matrice_dernier_prix(historique, col)),
                    (f"Positionnement {suffixe}", positionnement(historique,
                                                                 colonne=col)),
                ):
                    if tab is not None and not tab.empty:
                        tab.to_excel(xw, sheet_name=nom[:31])

            for nom, tab in (
                ("Volatilité", volatilite(historique)),
                ("Évolution Hympyr", evolution_periode(historique)),
                ("Dégressivité vrac", degressivite(historique)),
                ("Contrôle bases TVA", coherence_bases(historique)),
            ):
                if tab is not None and not tab.empty:
                    tab.to_excel(xw, sheet_name=nom[:31])

        modele_import().to_excel(xw, sheet_name="Modèle import", index=False)
        _formater_classeur(xw.book)
    return buf.getvalue()


class RapportPDF:
    """Rapport PDF à la charte Hympyr (fpdf2 + police Unicode embarquée)."""

    def __init__(self, titre: str, sous_titre: str = ""):
        from fpdf import FPDF

        self.pdf = FPDF(orientation="P", unit="mm", format="A4")
        self.pdf.set_auto_page_break(auto=True, margin=18)
        self._polices()
        self.titre, self.sous_titre = titre, sous_titre
        self.pdf.add_page()
        self._bandeau()

    def _polices(self):
        """Poppins si présente dans assets/, sinon DejaVu (fournie par matplotlib)."""
        try:
            base = os.path.join(os.path.dirname(os.path.abspath(__file__)), "assets")
        except NameError:
            base = "assets"
        pr = os.path.join(base, "Poppins-Regular.ttf")
        pb = os.path.join(base, "Poppins-Bold.ttf")
        if os.path.exists(pr) and os.path.exists(pb):
            self.pdf.add_font("Marque", "", pr)
            self.pdf.add_font("Marque", "B", pb)
        else:
            d = os.path.join(matplotlib.get_data_path(), "fonts", "ttf")
            self.pdf.add_font("Marque", "", os.path.join(d, "DejaVuSans.ttf"))
            self.pdf.add_font("Marque", "B", os.path.join(d, "DejaVuSans-Bold.ttf"))
        self.pdf.set_font("Marque", "", 10)

    @staticmethod
    def _rgb(hexa):
        h = hexa.lstrip("#")
        return tuple(int(h[i:i + 2], 16) for i in (0, 2, 4))

    def _bandeau(self):
        p = self.pdf
        p.set_fill_color(*self._rgb(VERT_FONCE))
        p.rect(0, 0, 210, 30, "F")
        p.set_text_color(255, 255, 255)
        p.set_xy(14, 8)
        p.set_font("Marque", "B", 14)
        p.cell(0, 8, self.titre, new_x="LMARGIN", new_y="NEXT")
        p.set_x(14)
        p.set_font("Marque", "", 8.5)
        p.cell(0, 6, self.sous_titre, new_x="LMARGIN", new_y="NEXT")
        p.set_text_color(40, 40, 40)
        p.set_y(40)

    def titre_section(self, txt):
        p = self.pdf
        if p.get_y() > 245:
            p.add_page()
        p.ln(3)
        p.set_font("Marque", "B", 12)
        p.set_text_color(*self._rgb(VERT_FONCE))
        p.set_x(14)
        p.cell(0, 8, txt, new_x="LMARGIN", new_y="NEXT")
        p.set_draw_color(*self._rgb(VERT_VIF))
        p.set_line_width(0.6)
        y = p.get_y()
        p.line(14, y, 196, y)
        p.set_text_color(40, 40, 40)
        p.ln(4)

    def paragraphe(self, txt, taille=9.5):
        self.pdf.set_font("Marque", "", taille)
        self.pdf.set_x(14)
        self.pdf.multi_cell(182, 5, txt)
        self.pdf.ln(2)

    def _entete_tableau(self, cols, largeurs, taille):
        p = self.pdf
        p.set_font("Marque", "B", taille)
        p.set_fill_color(*self._rgb(VERT_FONCE))
        p.set_text_color(255, 255, 255)
        p.set_x(14)
        for c, w in zip(cols, largeurs):
            p.cell(w, 7, c[:30], border=0, align="C", fill=True)
        p.ln()
        p.set_text_color(40, 40, 40)
        p.set_font("Marque", "", taille)

    def tableau(self, df: pd.DataFrame, index_label: str = None,
                max_lignes: int = 28):
        p = self.pdf
        if df is None or df.empty:
            self.paragraphe("Aucune donnée disponible.")
            return

        d = df.head(max_lignes).copy()
        if index_label:
            d = d.reset_index()
            d = d.rename(columns={d.columns[0]: index_label})

        cols = [str(c) for c in d.columns]
        valeurs = [[self._fmt(v) for v in row] for row in d.itertuples(index=False)]

        poids = [max([len(c)] + [len(r[i]) for r in valeurs]) if valeurs else len(c)
                 for i, c in enumerate(cols)]
        total = sum(poids) or 1
        largeurs = [max(15.0, 182 * w / total) for w in poids]
        facteur = 182 / sum(largeurs)
        largeurs = [w * facteur for w in largeurs]

        taille = 8 if len(cols) <= 6 else 6.6
        self._entete_tableau(cols, largeurs, taille)

        for i, row in enumerate(valeurs):
            if p.get_y() > 258:
                p.add_page()
                self._entete_tableau(cols, largeurs, taille)
            p.set_fill_color(*((240, 250, 243) if i % 2 == 0 else (255, 255, 255)))
            p.set_x(14)
            for v, w in zip(row, largeurs):
                p.cell(w, 6, v[:28], border=0, align="C", fill=True)
            p.ln()
        p.ln(3)

        if len(df) > max_lignes:
            self.paragraphe(f"({len(df) - max_lignes} ligne(s) non affichée(s) — "
                            "voir l'export Excel.)", 8)

    @staticmethod
    def _fmt(v):
        if v is None:
            return "—"
        if isinstance(v, (pd.Timestamp, datetime, date)):
            return pd.Timestamp(v).strftime("%d/%m/%Y")
        if isinstance(v, (bool, np.bool_)):
            return "oui" if v else "non"
        if isinstance(v, (int, np.integer)):
            return str(int(v))
        if isinstance(v, (float, np.floating)):
            return "—" if math.isnan(float(v)) else dec(float(v), 2)
        return str(v)

    def image(self, png: bytes, largeur=182):
        p = self.pdf
        if p.get_y() > 195:
            p.add_page()
        p.image(io.BytesIO(png), x=14, w=largeur)
        p.ln(4)

    def pied(self, mentions):
        p = self.pdf
        p.ln(6)
        p.set_font("Marque", "", 7.5)
        p.set_text_color(130, 130, 130)
        p.set_x(14)
        p.multi_cell(182, 4, mentions)
        p.set_text_color(40, 40, 40)

    def bytes(self) -> bytes:
        return bytes(self.pdf.output())


def pdf_grille(date_grille, palettes: dict, vrac: dict, zones: list,
               taux_march: float, taux_livr: float, franchise: float,
               prix_km_ht: float, prix_km_ttc: float) -> bytes:
    r = RapportPDF(
        "Grille tarifaire granulés de bois",
        f"{SOCIETE} · Date d'effet : {pd.Timestamp(date_grille):%d/%m/%Y} · "
        f"Édité le {datetime.now():%d/%m/%Y}")

    r.titre_section("Granulés en palette")
    r.tableau(pd.DataFrame([{"Produit": k, "Prix HT (€)": p["ht"],
                             "TVA": taux_fmt(taux_march),
                             "Prix TTC (€)": p["ttc"]}
                            for k, p in palettes.items()]))

    r.titre_section("Granulés en vrac")
    r.tableau(pd.DataFrame([
        {"Palier": f"{t} T", "HT / tonne (€)": p["ht"], "TVA": taux_fmt(taux_march),
         "TTC / tonne (€)": p["ttc"], "Total TTC (€)": round(t * p["ttc"], 2)}
        for t, p in sorted(vrac.items())]))

    r.titre_section("Frais de livraison — palettes")
    r.paragraphe("Frais déterminés par la distance routière réelle entre le dépôt "
                 "de Saint-Sulpice-la-Pointe et l'adresse de livraison. À partir de "
                 "deux palettes, les frais sont facturés une seule fois.")
    r.tableau(libelles_zones(zones, taux_livr))

    r.titre_section("Frais de livraison — vrac")
    r.paragraphe(
        f"Livraison offerte jusqu'à {dec(franchise, 0)} km. Au-delà : "
        f"{dec(prix_km_ht)} € HT ({dec(prix_km_ttc)} € TTC) par kilomètre "
        "supplémentaire, calculé sur la distance routière réelle.")

    r.pied(f"{SOCIETE} · {TELEPHONE} · Document indicatif à usage interne. "
           f"TVA marchandise {taux_fmt(taux_march)}, TVA livraison "
           f"{taux_fmt(taux_livr)} — taux paramétrés dans l'outil et à confirmer "
           "auprès du service comptable. Ne vaut pas offre contractuelle. "
           "Tarifs susceptibles de modification sans préavis.")
    return r.bytes()


def pdf_analyse(historique: pd.DataFrame, ref_focus: str = None,
                colonne: str = "prix_ttc") -> bytes:
    d = enrichir(historique)
    base = _base_lbl(colonne)
    if d.empty:
        r = RapportPDF("Analyse tarifaire", SOCIETE)
        r.paragraphe("Aucune donnée à analyser.")
        return r.bytes()

    d1, d2 = d["date"].min(), d["date"].max()
    acteurs = sorted(d["acteur"].unique())
    concurrents = [a for a in acteurs if a.lower() != "hympyr"]

    r = RapportPDF(
        f"Analyse tarifaire — granulés de bois ({base})",
        f"{SOCIETE} · Période {d1:%d/%m/%Y} → {d2:%d/%m/%Y} · "
        f"Édité le {datetime.now():%d/%m/%Y}")

    r.titre_section("Périmètre")
    r.paragraphe(
        f"{len(d)} relevés de prix · {len(acteurs)} acteur(s) dont "
        f"{len(concurrents)} concurrent(s) · {d['reference'].nunique()} référence(s) · "
        f"période du {d1:%d/%m/%Y} au {d2:%d/%m/%Y}.\n"
        f"Acteurs : {', '.join(acteurs)}.\n"
        f"Toutes les comparaisons de ce rapport sont exprimées en {base}.")

    coh = coherence_bases(historique)
    if not coh.empty and bool(coh["Anomalie"].any()):
        r.titre_section("⚠ Contrôle de cohérence")
        r.paragraphe(
            "Des taux de TVA anormaux ont été détectés : soit hétérogènes au sein "
            "d'un même acteur, soit divergents du taux dominant du jeu de données. "
            "Cela peut signaler un mélange de sources HT et TTC, auquel cas les "
            "écarts de prix ci-dessous seraient artificiels. À vérifier avant "
            "toute décision.")
        r.tableau(coh[coh["Anomalie"]], index_label="Acteur")

    pos = positionnement(historique, colonne=colonne)
    if not pos.empty:
        r.titre_section(f"Positionnement concurrentiel — dernier prix connu ({base})")
        moy = pos["Écart % vs moyenne"].mean()
        sens = "en dessous de" if moy < 0 else "au-dessus de"
        r.paragraphe(
            f"Sur l'ensemble des références comparables, Hympyr se situe en moyenne "
            f"{dec(abs(moy), 1)} % {sens} la moyenne du marché. Indice moyen : "
            f"{dec(pos['Indice (base 100)'].mean(), 1)} (base 100 = marché).")
        r.tableau(pos[["Hympyr", "Moyenne marché", "Mini marché", "Maxi marché",
                       "Écart % vs moyenne", "Rang", "Nb offres comparées"]],
                  index_label="Référence")
        r.image(fig_to_png(fig_positionnement(pos, base)))

    ev = evolution_periode(historique, colonne=colonne)
    if not ev.empty and d["date"].nunique() > 1:
        r.titre_section(f"Évolution des prix Hympyr sur la période ({base})")
        r.tableau(ev, index_label="Référence")

    if ref_focus:
        r.titre_section(f"Focus référence — {ref_focus}")
        r.image(fig_to_png(fig_vs_marche(historique, ref_focus, colonne=colonne)))
        r.image(fig_to_png(fig_evolution(historique, ref_focus, colonne)))

    deg = degressivite(historique, colonne)
    if not deg.empty:
        r.titre_section(f"Dégressivité vrac ({base})")
        r.paragraphe("Comparaison des structures de dégressivité : une pente plus "
                     "forte signale une politique plus agressive sur les gros volumes.")
        r.image(fig_to_png(fig_degressivite(historique, colonne)))
        r.tableau(deg, index_label="Palier (T)")

    vol = volatilite(historique, colonne)
    if not vol.empty:
        r.titre_section(f"Volatilité par référence ({base})")
        r.paragraphe("Amplitude entre prix mini et maxi relevés sur la période, "
                     "tous acteurs confondus.")
        r.tableau(vol, index_label="Référence")

    r.pied(f"{SOCIETE} · {TELEPHONE} · Analyse fondée exclusivement sur les données "
           "importées par l'utilisateur ; leur exactitude et leur représentativité "
           "n'ont pas été vérifiées. Les montants complétés par conversion HT/TTC "
           "reposent sur les taux renseignés dans l'outil. Document interne.")
    return r.bytes()

# ══════════════════════════════════════════════════════════════════════════════
# ===== LOGIQUE PURE — FIN =====================================================
# ══════════════════════════════════════════════════════════════════════════════


# ── Géocodage / routage (cache Streamlit) ─────────────────────────────────────

@st.cache_data(ttl=60 * 60 * 24 * 30, show_spinner=False)
def geocoder_adresse(adresse: str):
    try:
        r = requests.get("https://api-adresse.data.gouv.fr/search/",
                         params={"q": adresse, "limit": 1}, timeout=8)
        r.raise_for_status()
        feats = r.json().get("features", [])
        if not feats:
            return None
        lon, lat = feats[0]["geometry"]["coordinates"]
        return (lat, lon, feats[0]["properties"].get("label", adresse))
    except Exception:
        return None


@st.cache_data(ttl=60 * 60 * 24 * 30, show_spinner=False)
def chercher_communes(code_postal: str, ville: str = ""):
    params = {"limit": 15, "type": "municipality"}
    if ville.strip():
        params["q"] = ville.strip()
        params["postcode"] = code_postal
    else:
        params["q"] = code_postal
        params["postcode"] = code_postal
    try:
        r = requests.get("https://api-adresse.data.gouv.fr/search/",
                         params=params, timeout=8)
        r.raise_for_status()
        out = []
        for f in r.json().get("features", []):
            p = f["properties"]
            lon, lat = f["geometry"]["coordinates"]
            out.append({"label": p.get("label", ""),
                        "ville": p.get("city") or p.get("name", ""),
                        "cp": p.get("postcode", ""), "lat": lat, "lon": lon})
        return out
    except Exception:
        return []


@st.cache_data(ttl=60 * 60 * 24 * 30, show_spinner=False)
def distance_routiere(lat1, lon1, lat2, lon2):
    try:
        url = (f"https://router.project-osrm.org/route/v1/driving/"
               f"{lon1},{lat1};{lon2},{lat2}")
        r = requests.get(url, params={"overview": "false", "alternatives": "false"},
                         timeout=10)
        r.raise_for_status()
        data = r.json()
        if data.get("code") == "Ok" and data.get("routes"):
            route = data["routes"][0]
            return (route["distance"] / 1000.0, route["duration"] / 60.0, "osrm")
    except Exception:
        pass
    km = haversine(lat1, lon1, lat2, lon2) * COEF_VOL_OISEAU
    return (km, km / 55.0 * 60.0, "estimation")


# ══════════════════════════════════════════════════════════════════════════════
# ÉTAT DE SESSION
# ══════════════════════════════════════════════════════════════════════════════

def init_state():
    ss = st.session_state
    ss.setdefault("tva_march", TVA_MARCHANDISE_DEFAUT)
    ss.setdefault("tva_livr", TVA_LIVRAISON_DEFAUT)
    ss.setdefault("base_saisie", "TTC")

    tm, tl = ss["tva_march"], ss["tva_livr"]
    ss.setdefault("palettes", {
        k: {"ht": ttc_vers_ht(v, tm), "ttc": v}
        for k, v in PALETTES_TTC_DEFAUT.items()})
    ss.setdefault("vrac", {
        t: {"ht": ttc_vers_ht(v, tm), "ttc": v}
        for t, v in VRAC_TTC_DEFAUT.items()})
    ss.setdefault("zones", [
        {"Zone": z["Zone"], "Jusqu'à (km)": z["Jusqu'à (km)"],
         "72 h HT": ttc_vers_ht(z["72 h TTC"], tl), "72 h TTC": z["72 h TTC"],
         "15 j HT": ttc_vers_ht(z["15 j TTC"], tl), "15 j TTC": z["15 j TTC"]}
        for z in ZONES_TTC_DEFAUT])
    ss.setdefault("franchise_km", VRAC_FRANCHISE_KM_DEFAUT)
    ss.setdefault("prix_km_ttc", VRAC_PRIX_KM_TTC_DEFAUT)
    ss.setdefault("prix_km_ht", ttc_vers_ht(VRAC_PRIX_KM_TTC_DEFAUT, tl))
    ss.setdefault("date_grille", date.today())
    ss.setdefault("historique", df_vide())


init_state()

# ══════════════════════════════════════════════════════════════════════════════
# STYLE
# ══════════════════════════════════════════════════════════════════════════════

st.markdown(f"""
<style>
@import url('https://fonts.googleapis.com/css2?family=Poppins:wght@400;500;600;700&display=swap');
html, body, [class*="css"] {{ font-family: 'Poppins', sans-serif; }}

.hympyr-header {{
    background: linear-gradient(135deg, {VERT_FONCE} 0%, {VERT_VIF} 100%);
    border-radius: 16px; padding: 24px 30px; margin-bottom: 20px; color: white;
}}
.hympyr-header h1 {{ font-size: 1.5rem; font-weight: 700; margin: 0; }}
.hympyr-header p  {{ margin: 6px 0 0; opacity: .85; font-size: .86rem; }}

.result-card {{
    border-radius: 14px; padding: 20px 24px; margin: 16px 0;
    border-left: 6px solid {VERT_VIF}; background: {VERT_CLAIR};
}}
.card-warn {{ background: #fffaf0; border-color: {ORANGE}; }}
.card-err  {{ background: #fff4f4; border-color: {ROUGE}; }}

.zone-badge {{
    display: inline-block; background: {VERT_VIF}; color: white;
    font-size: .95rem; font-weight: 700; padding: 4px 14px;
    border-radius: 20px; margin-bottom: 10px;
}}
.total-box {{
    background: linear-gradient(135deg, {VERT_FONCE} 0%, {VERT_VIF} 100%);
    color: white; border-radius: 14px; padding: 20px 24px;
    margin: 16px 0; text-align: center;
}}
.total-box .lbl {{ font-size: .8rem; opacity: .85; letter-spacing: 1px;
                   text-transform: uppercase; }}
.total-box .val {{ font-size: 2.3rem; font-weight: 700; line-height: 1.2; }}
.total-box .sub {{ font-size: .78rem; opacity: .8; }}

.ligne {{ display: flex; justify-content: space-between; padding: 9px 0;
          border-bottom: 1px dashed #d7ece0; font-size: .95rem; }}
.ligne:last-child {{ border-bottom: none; }}
.ligne .k {{ color: #444; }}
.ligne .v {{ font-weight: 600; color: {VERT_FONCE}; }}
.ligne .v small {{ font-weight: 400; color: #888; }}

div[data-testid="stTextInput"] input {{
    font-family: 'Poppins', sans-serif !important;
    border-radius: 10px !important; border: 2px solid #c8e6c9 !important;
}}
</style>
""", unsafe_allow_html=True)

st.markdown(f"""
<div class="hympyr-header">
    <h1>🌿 Cockpit granulés — {SOCIETE}</h1>
    <p>Départ : {ADRESSE_DEPART} · Distance routière réelle · Prix HT et TTC</p>
</div>
""", unsafe_allow_html=True)

# ══════════════════════════════════════════════════════════════════════════════
# BARRE LATÉRALE
# ══════════════════════════════════════════════════════════════════════════════

with st.sidebar:
    st.markdown("### ⚙️ Paramètres")
    adresse_depart = st.text_input("Adresse de départ", value=ADRESSE_DEPART)
    aller_retour = st.checkbox(
        "Compter l'aller-retour", value=False,
        help="Décoché : seul le trajet dépôt → client est facturé.")
    forcer_km = st.checkbox("Forcer le kilométrage", value=False)
    km_manuel = st.number_input("Distance retenue (km)", 0.0, 500.0, 0.0, 1.0,
                                disabled=not forcer_km)
    st.divider()
    st.caption(f"TVA marchandise : **{taux_fmt(st.session_state['tva_march'])}** · "
               f"TVA livraison : **{taux_fmt(st.session_state['tva_livr'])}**  \n"
               "Modifiables dans l'onglet « Grille tarifaire ».")
    st.divider()
    _h = st.session_state["historique"]
    st.metric("Relevés de prix en session", len(_h))
    if not _h.empty:
        _d = pd.to_datetime(_h["date"], errors="coerce").dropna()
        if not _d.empty:
            st.caption(f"{_h['acteur'].nunique()} acteur(s) · "
                       f"{_d.min():%d/%m/%y} → {_d.max():%d/%m/%y}")
    st.caption("⚠️ Les données ne survivent pas à la fermeture de l'onglet. "
               "Exportez le classeur Excel pour archiver, réimportez-le pour reprendre.")

ong1, ong2, ong3, ong4 = st.tabs([
    "🧮 Simulateur", "📋 Grille tarifaire", "📥 Données", "📊 Analyse"])

# ══════════════════════════════════════════════════════════════════════════════
# ONGLET 1 — SIMULATEUR
# ══════════════════════════════════════════════════════════════════════════════

with ong1:
    tva_m = st.session_state["tva_march"]
    tva_l = st.session_state["tva_livr"]

    st.markdown("#### 📍 Adresse de livraison")
    c1, c2 = st.columns([1, 2])
    with c1:
        cp = st.text_input("Code postal", placeholder="81100", max_chars=5)
    with c2:
        ville = st.text_input("Ville (recommandé)", placeholder="Castres")

    destination = km_reel = minutes = source = None

    if cp and cp.strip().isdigit() and len(cp.strip()) == 5:
        communes = chercher_communes(cp.strip(), ville)
        if not communes:
            st.markdown(
                f"""<div class="result-card card-err">
                <strong>❌ Adresse introuvable</strong><br>
                <span style="font-size:.88rem;color:#666;">
                Aucune commune ne correspond au code postal {cp}. Vérifiez la saisie
                ou forcez le kilométrage dans le menu latéral.</span></div>""",
                unsafe_allow_html=True)
        elif len(communes) == 1:
            destination = communes[0]
            st.success(f"Destination : {destination['label']}")
        else:
            idx = st.selectbox("Plusieurs communes correspondent :",
                               range(len(communes)),
                               format_func=lambda i: communes[i]["label"])
            destination = communes[idx]

    if destination is not None:
        depart = geocoder_adresse(adresse_depart)
        if depart is None:
            st.error("Impossible de géocoder l'adresse de départ. "
                     "Forcez le kilométrage manuellement.")
        else:
            with st.spinner("Calcul de l'itinéraire…"):
                km_reel, minutes, source = distance_routiere(
                    depart[0], depart[1], destination["lat"], destination["lon"])

    km_retenu = None
    if forcer_km and km_manuel > 0:
        km_retenu, source = km_manuel, "manuel"
    elif km_reel is not None:
        km_retenu = km_reel * (2 if aller_retour else 1)

    if km_retenu is not None:
        km_retenu = arrondir_km(km_retenu)
        lib_source = {
            "osrm": "🛣️ Itinéraire routier réel",
            "estimation": "⚠️ Estimation (routeur indisponible) — à vérifier",
            "manuel": "✏️ Kilométrage saisi manuellement",
        }[source]
        trajet = "aller-retour" if (aller_retour and source != "manuel") else "aller simple"
        duree = f" · ~{minutes:.0f} min" if (minutes and source == "osrm") else ""
        st.markdown(
            f"""<div class="result-card {'card-warn' if source == 'estimation' else ''}">
            <div class="zone-badge">{km_fmt(km_retenu)} — {trajet}</div>
            <div style="font-size:.85rem;color:#555;">{lib_source}{duree}</div>
            </div>""", unsafe_allow_html=True)

        if km_retenu > SEUIL_ALERTE_KM:
            st.warning(f"Distance supérieure à {SEUIL_ALERTE_KM:.0f} km : la grille "
                       "zone 5 s'applique, mais faites valider faisabilité et marge "
                       "par l'exploitation avant engagement.")

    st.markdown("#### 📦 Commande")
    mode = st.radio("Conditionnement", ["Palettes", "Vrac"], horizontal=True,
                    label_visibility="collapsed")

    lignes = []          # (libellé, ht, ttc)
    produit_ht = produit_ttc = 0.0
    frais_ht = frais_ttc = None
    detail_frais, tonnes = "", 0.0
    palettes = st.session_state["palettes"]
    vrac = st.session_state["vrac"]

    if mode == "Palettes":
        noms = list(palettes.keys())
        cols = st.columns(max(2, len(noms)))
        quantites = {}
        for nom, col in zip(noms, cols):
            with col:
                quantites[nom] = st.number_input(f"Palettes {nom}", 0, 20, 0, 1,
                                                 key=f"q_{nom}")
        delai = st.radio("Délai de livraison",
                         ["Sous 72 h (prioritaire)",
                          "Sous 15 jours (tournée optimisée)"])
        nb_total = sum(quantites.values())
        for nom, q in quantites.items():
            if q:
                ht, ttc = q * palettes[nom]["ht"], q * palettes[nom]["ttc"]
                produit_ht += ht
                produit_ttc += ttc
                lignes.append((f"{q} × Palette {nom}", ht, ttc))
        tonnes = float(nb_total)  # 1 palette ≈ 1 tonne
        if km_retenu is not None and nb_total > 0:
            nom_z, lib_z, z = zone_palette(km_retenu, st.session_state["zones"])
            rapide = delai.startswith("Sous 72")
            frais_ht = float(z["72 h HT"] if rapide else z["15 j HT"])
            frais_ttc = float(z["72 h TTC"] if rapide else z["15 j TTC"])
            detail_frais = (f"{nom_z} ({lib_z}) · "
                            f"{'sous 72 h' if rapide else 'sous 15 jours'}"
                            + (" · facturés une seule fois" if nb_total > 1 else ""))
    else:
        paliers = sorted(vrac.keys())
        tonnage = st.select_slider("Tonnage commandé", options=paliers,
                                   value=paliers[min(2, len(paliers) - 1)],
                                   format_func=lambda t: f"{t} T")
        p = vrac[tonnage]
        produit_ht, produit_ttc = tonnage * p["ht"], tonnage * p["ttc"]
        tonnes = float(tonnage)
        lignes.append((f"{tonnage} T vrac × {eur(p['ttc'])} TTC/T",
                       produit_ht, produit_ttc))
        if km_retenu is not None:
            fr = st.session_state["franchise_km"]
            frais_ht = frais_vrac(km_retenu, fr, st.session_state["prix_km_ht"])
            frais_ttc = frais_vrac(km_retenu, fr, st.session_state["prix_km_ttc"])
            excedent = max(0.0, km_retenu - fr)
            detail_frais = (
                f"Dans la franchise de {dec(fr, 0)} km — offerts" if excedent == 0
                else f"{km_fmt(excedent)} au-delà de {dec(fr, 0)} km × "
                     f"{dec(st.session_state['prix_km_ttc'])} € TTC/km")

    if produit_ttc > 0:
        st.markdown("#### 🧮 Devis")
        html = "".join(
            f'<div class="ligne"><span class="k">{k}</span>'
            f'<span class="v">{eur(t)}<br><small>{eur(h)} HT</small></span></div>'
            for k, h, t in lignes)
        if frais_ttc is not None:
            html += (f'<div class="ligne"><span class="k">Frais de livraison<br>'
                     f'<span style="font-size:.78rem;color:#888;">{detail_frais}'
                     f'</span></span><span class="v">{eur(frais_ttc)}<br>'
                     f'<small>{eur(frais_ht)} HT</small></span></div>')
        st.markdown(f'<div class="result-card">{html}</div>', unsafe_allow_html=True)

        if frais_ttc is None:
            st.info("Saisissez le code postal pour obtenir les frais et le total.")
        else:
            total_ht = produit_ht + frais_ht
            total_ttc = produit_ttc + frais_ttc
            tva_totale = round(total_ttc - total_ht, 2)
            st.markdown(
                f"""<div class="total-box">
                <div class="lbl">Total TTC</div>
                <div class="val">{eur(total_ttc)}</div>
                <div class="sub">{eur(total_ht)} HT + {eur(tva_totale)} de TVA ·
                coût rendu : {eur(cout_livre(produit_ttc, frais_ttc, tonnes))}
                TTC / tonne</div>
                </div>""", unsafe_allow_html=True)

            k1, k2, k3 = st.columns(3)
            k1.metric("Total HT", eur(total_ht))
            k2.metric("TVA", eur(tva_totale),
                      help=f"Marchandise {taux_fmt(tva_m)} · "
                           f"Livraison {taux_fmt(tva_l)}")
            k3.metric("Total TTC", eur(total_ttc))

            dest_txt = destination["label"] if destination else f"{cp} {ville}".strip()
            recap = [f"Devis granulés — {datetime.now():%d/%m/%Y}",
                     f"Livraison : {dest_txt}",
                     f"Distance : {km_fmt(km_retenu)} "
                     f"({'aller-retour' if aller_retour else 'aller simple'})", ""]
            recap += [f"- {k} : {eur(t)} TTC ({eur(h)} HT)" for k, h, t in lignes]
            recap += [f"- Frais de livraison ({detail_frais}) : "
                      f"{eur(frais_ttc)} TTC ({eur(frais_ht)} HT)", "",
                      f"TOTAL HT   : {eur(total_ht)}",
                      f"TVA        : {eur(tva_totale)}",
                      f"TOTAL TTC  : {eur(total_ttc)}"]
            with st.expander("📋 Récapitulatif à copier (mail / téléphone)"):
                st.code("\n".join(recap), language=None)

# ══════════════════════════════════════════════════════════════════════════════
# ONGLET 2 — GRILLE TARIFAIRE
# ══════════════════════════════════════════════════════════════════════════════

with ong2:
    st.markdown("#### 📋 Grille tarifaire en vigueur")
    st.caption("Les prix validés ici alimentent immédiatement le simulateur. "
               "Plus besoin de recoder lors d'une hausse fournisseur.")

    t1, t2, t3 = st.columns([1, 1, 1])
    with t1:
        d_grille = st.date_input("Date d'effet", value=st.session_state["date_grille"],
                                 format="DD/MM/YYYY")
        st.session_state["date_grille"] = d_grille
    with t2:
        tva_m_pct = st.number_input(
            "TVA marchandise (%)", 0.0, 30.0,
            float(st.session_state["tva_march"] * 100), 0.1, format="%.1f")
    with t3:
        tva_l_pct = st.number_input(
            "TVA livraison (%)", 0.0, 30.0,
            float(st.session_state["tva_livr"] * 100), 0.1, format="%.1f")

    tva_m_saisi, tva_l_saisi = tva_m_pct / 100.0, tva_l_pct / 100.0

    st.warning("⚠️ Les taux de TVA sont des paramètres de l'outil, pas une "
               "recommandation fiscale. Faites-les valider par la comptabilité — "
               "le taux applicable dépend du produit, de l'usage et du statut du "
               "client, et le régime des frais de port doit être confirmé.")

    base = st.radio(
        "Base de saisie des montants", ["TTC", "HT"], horizontal=True,
        index=0 if st.session_state["base_saisie"] == "TTC" else 1,
        help="La colonne saisie fait foi ; l'autre est calculée puis figée à la "
             "validation. Aucun aller-retour de conversion, donc aucune dérive "
             "d'arrondi.")
    autre = "HT" if base == "TTC" else "TTC"
    cle, cle_autre = base.lower(), autre.lower()

    ca, cb = st.columns(2)
    with ca:
        st.markdown(f"**Palettes** — € {base} / palette")
        dfp = st.data_editor(
            pd.DataFrame([{"Produit": k, f"Prix {base}": float(v[cle]),
                           f"Prix {autre} (calculé)": float(v[cle_autre])}
                          for k, v in st.session_state["palettes"].items()]),
            num_rows="dynamic", use_container_width=True, hide_index=True,
            disabled=[f"Prix {autre} (calculé)"], key="ed_palettes")
    with cb:
        st.markdown(f"**Vrac** — € {base} / tonne par palier")
        dfv = st.data_editor(
            pd.DataFrame([{"Palier (T)": int(t), f"Prix {base} / T": float(v[cle]),
                           f"Prix {autre} / T (calculé)": float(v[cle_autre])}
                          for t, v in sorted(st.session_state["vrac"].items())]),
            num_rows="dynamic", use_container_width=True, hide_index=True,
            disabled=[f"Prix {autre} / T (calculé)"], key="ed_vrac")

    st.markdown(f"**Frais de livraison palettes** — € {base}, bornes continues en km")
    dfz = st.data_editor(
        pd.DataFrame([{"Zone": z["Zone"], "Jusqu'à (km)": float(z["Jusqu'à (km)"]),
                       f"72 h {base}": float(z[f"72 h {base}"]),
                       f"15 j {base}": float(z[f"15 j {base}"])}
                      for z in st.session_state["zones"]]),
        num_rows="dynamic", use_container_width=True, hide_index=True, key="ed_zones")

    v1, v2 = st.columns(2)
    with v1:
        franchise_saisie = st.number_input(
            "Vrac — franchise de livraison (km)", 0.0, 200.0,
            float(st.session_state["franchise_km"]), 1.0)
    with v2:
        prix_km_saisi = st.number_input(
            f"Vrac — € {base} par km au-delà", 0.0, 20.0,
            float(st.session_state["prix_km_ttc"] if base == "TTC"
                  else st.session_state["prix_km_ht"]), 0.01, format="%.2f")

    if st.button("💾 Appliquer les modifications", type="primary"):
        try:
            npal = {}
            for _, r in dfp.iterrows():
                nom = str(r["Produit"]).strip()
                val = r[f"Prix {base}"]
                if not nom or pd.isna(val):
                    continue
                ht, ttc = paire_prix(val, tva_m_saisi, base)
                npal[nom] = {"ht": ht, "ttc": ttc}

            nvrac = {}
            for _, r in dfv.iterrows():
                if pd.isna(r["Palier (T)"]) or pd.isna(r[f"Prix {base} / T"]):
                    continue
                ht, ttc = paire_prix(r[f"Prix {base} / T"], tva_m_saisi, base)
                nvrac[int(r["Palier (T)"])] = {"ht": ht, "ttc": ttc}

            nzones = []
            for _, r in dfz.iterrows():
                if pd.isna(r["Jusqu'à (km)"]) or pd.isna(r[f"72 h {base}"]) \
                        or pd.isna(r[f"15 j {base}"]):
                    continue
                h72, t72 = paire_prix(r[f"72 h {base}"], tva_l_saisi, base)
                h15, t15 = paire_prix(r[f"15 j {base}"], tva_l_saisi, base)
                nzones.append({
                    "Zone": str(r["Zone"]).strip() or "Zone",
                    "Jusqu'à (km)": float(r["Jusqu'à (km)"]),
                    "72 h HT": h72, "72 h TTC": t72,
                    "15 j HT": h15, "15 j TTC": t15})

            if not npal or not nvrac or not nzones:
                st.error("Chaque tableau doit contenir au moins une ligne valide.")
            else:
                km_ht, km_ttc = paire_prix(prix_km_saisi, tva_l_saisi, base)
                st.session_state.update({
                    "palettes": npal, "vrac": nvrac, "zones": nzones,
                    "tva_march": tva_m_saisi, "tva_livr": tva_l_saisi,
                    "base_saisie": base, "franchise_km": float(franchise_saisie),
                    "prix_km_ht": km_ht, "prix_km_ttc": km_ttc})
                st.success(f"Grille mise à jour (saisie en {base}, TVA marchandise "
                           f"{taux_fmt(tva_m_saisi)}, TVA livraison "
                           f"{taux_fmt(tva_l_saisi)}). Le simulateur utilise ces "
                           "valeurs.")
                st.rerun()
        except Exception as e:
            st.error(f"Saisie invalide : {e}")

    st.divider()
    st.markdown("**Aperçu de la grille validée**")
    st.dataframe(grille_tableau(st.session_state["palettes"],
                                st.session_state["vrac"],
                                st.session_state["tva_march"]),
                 use_container_width=True, hide_index=True)

    st.divider()
    _args = (st.session_state["date_grille"], st.session_state["palettes"],
             st.session_state["vrac"], st.session_state["zones"],
             st.session_state["tva_march"], st.session_state["tva_livr"],
             st.session_state["franchise_km"], st.session_state["prix_km_ht"],
             st.session_state["prix_km_ttc"])

    e1, e2, e3 = st.columns(3)
    with e1:
        st.download_button(
            "📄 Exporter la grille (PDF)", data=pdf_grille(*_args),
            file_name=f"HYM_grille_granules_"
                      f"{pd.Timestamp(st.session_state['date_grille']):%Y%m%d}.pdf",
            mime="application/pdf", use_container_width=True)
    with e2:
        st.download_button(
            "📊 Exporter le classeur (Excel)",
            data=export_excel(st.session_state["date_grille"],
                              st.session_state["palettes"], st.session_state["vrac"],
                              st.session_state["zones"], st.session_state["historique"],
                              st.session_state["tva_march"],
                              st.session_state["tva_livr"],
                              st.session_state["franchise_km"],
                              st.session_state["prix_km_ht"],
                              st.session_state["prix_km_ttc"]),
            file_name=f"HYM_grille_granules_"
                      f"{pd.Timestamp(st.session_state['date_grille']):%Y%m%d}.xlsx",
            mime=("application/vnd.openxmlformats-officedocument"
                  ".spreadsheetml.sheet"), use_container_width=True)
    with e3:
        if st.button("📌 Archiver cette grille dans l'historique",
                     use_container_width=True):
            nouv = grille_vers_lignes(st.session_state["date_grille"],
                                      st.session_state["palettes"],
                                      st.session_state["vrac"],
                                      st.session_state["tva_march"])
            st.session_state["historique"] = dedoublonner(
                pd.concat([st.session_state["historique"], nouv], ignore_index=True),
                st.session_state["tva_march"])
            st.success(f"{len(nouv)} lignes archivées au "
                       f"{pd.Timestamp(st.session_state['date_grille']):%d/%m/%Y}.")

# ══════════════════════════════════════════════════════════════════════════════
# ONGLET 3 — DONNÉES
# ══════════════════════════════════════════════════════════════════════════════

with ong3:
    tva_ref = st.session_state["tva_march"]

    st.markdown("#### 📥 Import de grilles tarifaires")
    st.caption("Formats acceptés : .xlsx, .csv — plusieurs fichiers à la fois. "
               "Colonnes minimales : date, acteur, et un prix (HT ou TTC). "
               "Le montant manquant est complété avec le taux de TVA de la ligne, "
               f"ou {taux_fmt(tva_ref)} à défaut.")

    fichiers = st.file_uploader("Fichiers à importer", type=["xlsx", "xls", "csv"],
                                accept_multiple_files=True)

    if fichiers and st.button("⬆️ Importer", type="primary"):
        total_ok, rejets, avert = 0, [], []
        for f in fichiers:
            try:
                brut = (pd.read_csv(f, sep=None, engine="python")
                        if f.name.lower().endswith(".csv") else pd.read_excel(f))
            except Exception as e:
                rejets.append(f"{f.name} : illisible ({e}).")
                continue
            ok, rj, av = valider_import(brut, f.name, tva_ref)
            rejets += rj
            avert += av
            if not ok.empty:
                st.session_state["historique"] = pd.concat(
                    [st.session_state["historique"], ok], ignore_index=True)
                total_ok += len(ok)

        if total_ok:
            st.session_state["historique"] = dedoublonner(
                st.session_state["historique"], tva_ref)
            st.success(f"{total_ok} ligne(s) importée(s).")
        for a in avert:
            st.info("ℹ️ " + a)
        for r in rejets:
            st.warning("⚠️ " + r)

    with st.expander("📎 Modèle d'import à distribuer"):
        st.dataframe(modele_import(), use_container_width=True, hide_index=True)
        st.caption("Renseigner prix_ht ou prix_ttc — ou les deux. Le taux de TVA "
                   "s'exprime indifféremment en 10, 10 % ou 0,10.")
        _buf = io.BytesIO()
        with pd.ExcelWriter(_buf, engine="openpyxl") as _xw:
            modele_import().to_excel(_xw, sheet_name="Modèle import", index=False)
        st.download_button("Télécharger le modèle (Excel)", data=_buf.getvalue(),
                           file_name="HYM_modele_import_prix.xlsx",
                           mime=("application/vnd.openxmlformats-officedocument"
                                 ".spreadsheetml.sheet"))

    st.divider()
    st.markdown("#### ✍️ Saisie directe d'un relevé concurrent")
    st.caption("Pour un prix relevé au téléphone ou sur un site. Précisez bien la "
               "base : un tarif affiché en ligne est généralement TTC, un devis "
               "professionnel généralement HT.")

    with st.form("saisie_conc"):
        f1, f2, f3 = st.columns(3)
        with f1:
            s_date = st.date_input("Date du relevé", value=date.today(),
                                   format="DD/MM/YYYY")
            s_acteur = st.text_input("Concurrent", placeholder="ex : Péchavy")
        with f2:
            s_cond = st.selectbox("Conditionnement", ["Palette", "Vrac"])
            s_produit = st.text_input("Produit", value="Granulés")
        with f3:
            s_palier = st.number_input("Palier (T) — vrac uniquement", 0, 30, 0, 1)
            s_base = st.radio("Base du prix relevé", ["TTC", "HT"], horizontal=True)
        g1, g2 = st.columns(2)
        with g1:
            s_prix = st.number_input(f"Prix {s_base}", 0.0, 5000.0, 0.0, 0.01)
        with g2:
            s_tva = st.number_input("TVA applicable (%)", 0.0, 30.0,
                                    float(tva_ref * 100), 0.1, format="%.1f")
        if st.form_submit_button("Ajouter le relevé", type="primary"):
            if not s_acteur.strip() or s_prix <= 0:
                st.error("Nom du concurrent et prix strictement positif obligatoires.")
            else:
                taux = s_tva / 100.0
                ht, ttc = paire_prix(s_prix, taux, s_base)
                ligne = pd.DataFrame([{
                    "date": pd.Timestamp(s_date), "acteur": s_acteur.strip(),
                    "type_acteur": "Concurrent", "conditionnement": s_cond,
                    "produit": s_produit.strip() or "Granulés",
                    "palier_t": (float(s_palier)
                                 if (s_cond == "Vrac" and s_palier) else np.nan),
                    "unite": "tonne" if s_cond == "Vrac" else "palette",
                    "prix_ht": ht, "taux_tva": taux, "prix_ttc": ttc,
                }], columns=COLONNES)
                st.session_state["historique"] = dedoublonner(
                    pd.concat([st.session_state["historique"], ligne],
                              ignore_index=True), tva_ref)
                st.success(f"Relevé {s_acteur} ajouté : {eur(ht)} HT / "
                           f"{eur(ttc)} TTC.")

    st.divider()
    st.markdown("#### 🗂️ Données en session")
    hist = st.session_state["historique"]
    if hist.empty:
        st.info("Aucune donnée. Importez un fichier, saisissez un relevé, ou "
                "archivez la grille depuis l'onglet « Grille tarifaire ».")
    else:
        aff = enrichir(hist, tva_ref).copy()
        aff["date"] = aff["date"].dt.strftime("%d/%m/%Y")
        aff["taux_tva"] = aff["taux_tva"].apply(taux_fmt)
        st.dataframe(aff, use_container_width=True, hide_index=True, height=340)

        coh = coherence_bases(hist)
        if not coh.empty and bool(coh["Anomalie"].any()):
            st.warning("Taux de TVA anormaux détectés (hétérogènes chez un acteur, "
                       "ou divergents du taux dominant) : vérifiez que vous ne "
                       "mélangez pas des relevés HT et TTC, ce qui fausserait les "
                       "écarts calculés dans l'analyse.")
            st.dataframe(coh[coh["Anomalie"]], use_container_width=True)

        if st.button("🗑️ Vider les données"):
            st.session_state["historique"] = df_vide()
            st.rerun()

# ══════════════════════════════════════════════════════════════════════════════
# ONGLET 4 — ANALYSE
# ══════════════════════════════════════════════════════════════════════════════

with ong4:
    hist = st.session_state["historique"]
    d = enrichir(hist, st.session_state["tva_march"])
    if d.empty:
        st.info("Aucune donnée à analyser. Rendez-vous dans l'onglet « Données ».")
    else:
        base_an = st.radio(
            "Base de comparaison", ["TTC", "HT"], horizontal=True,
            help="Comparez en HT face à des concurrents professionnels, en TTC "
                 "face à des tarifs grand public. Ne mélangez jamais les deux.")
        col_an = "prix_ttc" if base_an == "TTC" else "prix_ht"

        acteurs = sorted(d["acteur"].unique())
        concurrents = [a for a in acteurs if a.lower() != "hympyr"]
        refs = sorted(d["reference"].unique())

        m1, m2, m3, m4 = st.columns(4)
        m1.metric("Relevés", len(d))
        m2.metric("Acteurs", len(acteurs))
        m3.metric("Références", len(refs))
        m4.metric("Période", f"{d['date'].min():%m/%y} → {d['date'].max():%m/%y}")

        coh = coherence_bases(hist)
        if not coh.empty and bool(coh["Anomalie"].any()):
            st.error("⚠️ Taux de TVA anormaux détectés : les écarts "
                     "ci-dessous peuvent être artificiels. Corrigez les relevés "
                     "dans l'onglet « Données » avant d'exploiter ces chiffres.")

        if not concurrents:
            st.warning("Aucun concurrent renseigné : les lectures concurrentielles "
                       "resteront vides. Ajoutez des relevés dans l'onglet « Données ».")

        st.divider()
        st.markdown(f"##### 1. Matrice des prix — dernier prix connu ({base_an})")
        st.caption("Lecture transversale : qui pratique quoi, sur quelle référence.")
        mat = matrice_dernier_prix(hist, col_an)
        if mat.empty:
            st.info("Données insuffisantes.")
        else:
            sty = mat.style.format(lambda v: dec(v) if pd.notna(v) else "—")
            if mat.shape[1] > 1:
                sty = sty.background_gradient(cmap="RdYlGn_r", axis=1)
            st.dataframe(sty, use_container_width=True)

        st.divider()
        st.markdown(f"##### 2. Positionnement concurrentiel ({base_an})")
        st.caption("Indice base 100 = moyenne du marché. Inférieur à 100 : "
                   "Hympyr est moins cher.")
        pos = positionnement(hist, colonne=col_an)
        if pos.empty:
            st.info("Nécessite au moins un concurrent sur une référence commune "
                    "avec Hympyr.")
        else:
            st.dataframe(pos.style.format({
                "Hympyr": dec, "Moyenne marché": dec, "Mini marché": dec,
                "Maxi marché": dec, "Écart € vs moyenne": dec,
                "Écart % vs moyenne": pct,
                "Indice (base 100)": lambda v: dec(v, 1)}),
                use_container_width=True)
            st.pyplot(fig_positionnement(pos, base_an))
            moy = pos["Écart % vs moyenne"].mean()
            if moy < -3:
                st.success(f"Hympyr est en moyenne {dec(abs(moy), 1)} % sous le "
                           "marché. Vérifier que cet écart est un choix assumé, "
                           "pas une érosion de marge.")
            elif moy > 3:
                st.warning(f"Hympyr est en moyenne {dec(moy, 1)} % au-dessus du "
                           "marché. L'écart doit être argumenté par les commerciaux "
                           "(service, délai, proximité).")
            else:
                st.info(f"Hympyr est aligné sur le marché ({pct(moy)}).")

        st.divider()
        st.markdown(f"##### 3. Évolution dans le temps ({base_an})")
        ref_focus = st.selectbox("Référence à examiner", refs)
        g1, g2 = st.columns(2)
        with g1:
            st.pyplot(fig_evolution(hist, ref_focus, col_an))
        with g2:
            st.pyplot(fig_vs_marche(hist, ref_focus, colonne=col_an))

        ev = evolution_periode(hist, colonne=col_an)
        if not ev.empty:
            st.markdown(f"**Variation des prix Hympyr sur la période ({base_an})**")
            st.dataframe(ev.style.format({
                "Prix initial": dec, "Prix actuel": dec,
                "Variation €": dec, "Variation %": pct}),
                use_container_width=True)

        st.divider()
        st.markdown(f"##### 4. Dégressivité vrac ({base_an})")
        st.caption("Une pente plus forte traduit une politique plus agressive "
                   "sur les gros volumes.")
        deg = degressivite(hist, col_an)
        if deg.empty:
            st.info("Aucune donnée vrac avec palier renseigné.")
        else:
            st.pyplot(fig_degressivite(hist, col_an))
            st.dataframe(deg.style.format(lambda v: dec(v) if pd.notna(v) else "—"),
                         use_container_width=True)

        st.divider()
        st.markdown(f"##### 5. Volatilité par référence ({base_an})")
        st.dataframe(volatilite(hist, col_an).style.format({
            "Prix mini": dec, "Prix maxi": dec, "Prix moyen": dec,
            "Écart-type": dec, "Amplitude %": lambda v: pct(v, signe=False)}),
            use_container_width=True)

        st.divider()
        st.markdown("##### 📤 Export du rapport")
        r1, r2 = st.columns(2)
        with r1:
            st.download_button(
                f"📄 Rapport d'analyse {base_an} (PDF)",
                data=pdf_analyse(hist, ref_focus, col_an),
                file_name=f"HYM_analyse_prix_{base_an}_"
                          f"{datetime.now():%Y%m%d}.pdf",
                mime="application/pdf", use_container_width=True)
        with r2:
            st.download_button(
                "📊 Classeur complet (Excel)",
                data=export_excel(st.session_state["date_grille"],
                                  st.session_state["palettes"],
                                  st.session_state["vrac"],
                                  st.session_state["zones"], hist,
                                  st.session_state["tva_march"],
                                  st.session_state["tva_livr"],
                                  st.session_state["franchise_km"],
                                  st.session_state["prix_km_ht"],
                                  st.session_state["prix_km_ttc"]),
                file_name=f"HYM_analyse_prix_{datetime.now():%Y%m%d}.xlsx",
                mime=("application/vnd.openxmlformats-officedocument"
                      ".spreadsheetml.sheet"), use_container_width=True)

st.markdown(f"""
<div style="text-align:center;color:#aaa;font-size:.75rem;margin-top:30px;">
{SOCIETE} · Usage interne équipe granulés · {TELEPHONE}<br>
Simulations et analyses indicatives — ne valent pas engagement contractuel.
Taux de TVA paramétrés dans l'outil, à valider par la comptabilité.
</div>
""", unsafe_allow_html=True)
